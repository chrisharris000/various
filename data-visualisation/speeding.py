import openpyxl
import folium
from colour import Color
#bar graph

def retrieve_speeding_data(file_name):
    data_2014 = {}
    data_2015 = {}
    
    book = openpyxl.load_workbook(file_name)
    curr_sheet = book.get_sheet_by_name("Report ")

    for row in range(19,615):
        postcode = curr_sheet.cell(row = row, column = 2).value
        
        n_offences_speed_2014 = curr_sheet.cell(row = row, column = 3).value
        value_speed_2014 = curr_sheet.cell(row = row, column = 4).value
        n_offences_police_2014 = curr_sheet.cell(row = row, column = 7).value
        value_police_2014 = curr_sheet.cell(row = row, column = 8).value

        n_offences_speed_2015 = curr_sheet.cell(row = row, column = 5).value
        value_speed_2015 = curr_sheet.cell(row = row, column = 6).value
        n_offences_police_2015 = curr_sheet.cell(row = row, column = 9).value
        value_police_2015 = curr_sheet.cell(row = row, column = 10).value
        
        data_2014[postcode] = [[n_offences_speed_2014, value_speed_2014],[n_offences_police_2014, value_police_2014]]
        data_2015[postcode] = [[n_offences_speed_2015, value_speed_2015],[n_offences_police_2015, value_police_2015]]

    return curr_sheet, data_2014, data_2015

def retrieve_postocde_data(file_name):
    book = openpyxl.load_workbook(file_name)
    curr_sheet = book.get_sheet_by_name("Final Data")

    collated_data = {}
    metro_pcode = []
    
    
    for row in range(1,4768):
        postcode = curr_sheet.cell(row = row, column = 1).value
        suburb = curr_sheet.cell(row = row, column = 2).value
        lat = curr_sheet.cell(row = row, column = 4).value
        long = curr_sheet.cell(row = row, column = 5).value
        region = curr_sheet.cell(row = row, column = 6).value

        collated_data[suburb] = [postcode, lat, long, region]

        if region == "Metro" and postcode not in metro_pcode:
            metro_pcode.append(postcode)
        
    return collated_data,metro_pcode;
        

def retrieve_ranked_n(data,n,top = True):
    fine_rank = []
    pcode_map = {}
    top_n = []
    top_n_pcode = []
    bottom_n = []
    bottom_n_pcode = []
    
    for entry in data:
        fine_value = data[entry][0][1] + data[entry][1][1]
        fine_rank.append(fine_value)
        pcode_map[entry] = fine_value
        
    fine_rank = sorted(fine_rank)

    if top:
        for i in range(len(fine_rank)-1, len(fine_rank) - n - 1, -1):
            top_n.append(fine_rank[i])
            for entry in pcode_map:
                if pcode_map[entry] == fine_rank[i]:
                    top_n_pcode.append(entry)
        return top_n, top_n_pcode
    
    else:
        for i in range(n):
            bottom_n.append(fine_rank[i])
            for entry in pcode_map:
                if pcode_map[entry] == fine_rank[i]:
                    bottom_n_pcode.append(entry)
        return bottom_n, bottom_n_pcode
    
def query_specific_pcode(data,pcode):
    fine_rank = []
    pcode_map = {}
    
    for entry in data:
        fine_value = data[entry][0][1] + data[entry][1][1]
        fine_rank.append(fine_value)
        pcode_map[entry] = fine_value
        
    fine_rank = sorted(fine_rank)
    return pcode_map[pcode], len(fine_rank) - fine_rank.index(pcode_map[pcode])

def text_results(data_2014, data_2015, n):
    h_fine, h_pcode = retrieve_ranked_n(data_2014,n, top = True)
    print("Top",n,"Postcodes for Speeding Fines 2014")
    for p in range(len(h_pcode)):
        if h_pcode[p] in metro_pcode:
            print(str(p + 1) + '.\t', h_pcode[p], "$" + str(h_fine[p]) + ' - Metro')
        else:
            print(str(p + 1) + '.\t', h_pcode[p], "$" + str(h_fine[p]) + ' - Regional')

    h_fine, h_pcode = retrieve_ranked_n(data_2015,n, top = True)
    print("\nTop",n,"Postcodes for Speeding Fines 2015")
    for p in range(len(h_pcode)):
        if h_pcode[p] in metro_pcode:
            print(str(p + 1) + '.\t', h_pcode[p], "$" + str(h_fine[p]) + ' - Metro')
        else:
            print(str(p + 1) + '.\t', h_pcode[p], "$" + str(h_fine[p]) + ' - Regional')

    h_fine, h_pcode = retrieve_ranked_n(data_2014,n, top = False)
    print("\nBottom",n,"Postcodes for Speeding Fines 2014")
    for p in range(len(h_pcode)):
        if h_pcode[p] in metro_pcode:
            print(str(p + 1) + '.\t', h_pcode[p], "$" + str(h_fine[p]) + ' - Metro')
        else:
            print(str(p + 1) + '.\t', h_pcode[p], "$" + str(h_fine[p]) + ' - Regional')   

    h_fine, h_pcode = retrieve_ranked_n(data_2015,n, top = False)
    print("\nBottom",n,"Postcodes for Speeding Fines 2015")
    for p in range(len(h_pcode)):
        if h_pcode[p] in metro_pcode:
            print(str(p + 1) + '.\t', h_pcode[p], "$" + str(h_fine[p]) + ' - Metro')
        else:
            print(str(p + 1) + '.\t', h_pcode[p], "$" + str(h_fine[p]) + ' - Regional')

'''def create_heat_map(speeding_data, postcode_data):
    #tried implementing but doesn't work as intended
    latitudes = []
    longitudes = []
    
    for entry in postcode_data:
        pcode = postcode_data[entry][0]
        if pcode in speeding_data:
            fine_total = speeding_data[pcode][0][0] + speeding_data[pcode][1][0]
            mod_fine_total = fine_total % 100
            for i in range(mod_fine_total):
                latitudes.append(postcode_data[entry][1])
                longitudes.append(postcode_data[entry][2])

    gmap = gmplot.GoogleMapPlotter(-33.8688, 151.2093, 10)
    gmap.heatmap(latitudes, longitudes)
    gmap.draw("test.html")'''

def circle_ranked_n(speeding_data, postcode_data):   
    h_fine, h_pcode = retrieve_ranked_n(speeding_data,10,top = True)
    latitude = []
    longitude = []
    red = Color("red")
    orange = Color("Orange")
    colors = list(red.range_to(orange,10))
    map_osm = folium.Map(location=[-33.8688, 151.2093],zoom_start = 13)
    
    for h_p in h_pcode:
        avg_lat = 0
        avg_lon = 0
        total = 0
        for entry in postcode_data:
            if postcode_data[entry][0] == h_p:
                avg_lat += postcode_data[entry][1]
                avg_lon += postcode_data[entry][2]
                total += 1
        msg = str(h_p) + " - Ranked " + str(h_pcode.index(h_p)+1) +" for Speeding Fines"
        c = colors[h_pcode.index(h_p)]
        folium.CircleMarker([avg_lat/total,avg_lon/total],
                    radius=75,
                    popup=msg,
                    color=c.hex,
                    fill_color=c.hex,
                    fill = True
                   ).add_to(map_osm)   
                
    map_osm.save("test.html")
    
curr_sheet, data_2014, data_2015 = retrieve_speeding_data("./speeding_stats.xlsx")
collated_data, metro_pcode = retrieve_postocde_data("./Australian_Post_Codes_Lat_Lon.xlsx")

circle_ranked_n(data_2014, collated_data)
