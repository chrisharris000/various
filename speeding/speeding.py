import openpyxl
import folium
from folium import plugins
from colour import Color
import string

def retrieve_speeding_data(file_name):
    data_2014 = {}
    data_2015 = {}
    
    book = openpyxl.load_workbook(file_name)
    curr_sheet = book.get_sheet_by_name("Report ")

    for row in range(19,615):
        postcode = curr_sheet.cell(row = row, column = 2).value
        
        n_offences_speed_2014 = curr_sheet.cell(row = row, column = 3).value
        value_speed_2014 = curr_sheet.cell(row = row, column = 4).value
        n_offences_police_2014 = curr_sheet.cell(row = row, column = 7).value
        value_police_2014 = curr_sheet.cell(row = row, column = 8).value

        n_offences_speed_2015 = curr_sheet.cell(row = row, column = 5).value
        value_speed_2015 = curr_sheet.cell(row = row, column = 6).value
        n_offences_police_2015 = curr_sheet.cell(row = row, column = 9).value
        value_police_2015 = curr_sheet.cell(row = row, column = 10).value
        
        data_2014[postcode] = [[n_offences_speed_2014, value_speed_2014],[n_offences_police_2014, value_police_2014]]
        data_2015[postcode] = [[n_offences_speed_2015, value_speed_2015],[n_offences_police_2015, value_police_2015]]

    return curr_sheet, data_2014, data_2015

def retrieve_postocde_data(file_name):
    book = openpyxl.load_workbook(file_name)
    curr_sheet = book.get_sheet_by_name("Final Data")

    collated_data = {}
    metro_pcode = []
    
    
    for row in range(1,4768):
        postcode = curr_sheet.cell(row = row, column = 1).value
        suburb = curr_sheet.cell(row = row, column = 2).value
        lat = curr_sheet.cell(row = row, column = 4).value
        long = curr_sheet.cell(row = row, column = 5).value
        region = curr_sheet.cell(row = row, column = 6).value

        collated_data[suburb] = [postcode, lat, long, region]

        if region == "Metro" and postcode not in metro_pcode:
            metro_pcode.append(postcode)
        
    return collated_data,metro_pcode;
        

def retrieve_ranked_n(data,n,top = True):
    fine_rank = []
    pcode_map = {}
    top_n = []
    top_n_pcode = []
    bottom_n = []
    bottom_n_pcode = []
    
    for entry in data:
        fine_value = data[entry][0][1] + data[entry][1][1]
        fine_rank.append(fine_value)
        pcode_map[entry] = fine_value
        
    fine_rank = sorted(fine_rank)

    if top:
        for i in range(len(fine_rank)-1, len(fine_rank) - n - 1, -1):
            top_n.append(fine_rank[i])
            for entry in pcode_map:
                if pcode_map[entry] == fine_rank[i]:
                    top_n_pcode.append(entry)
        return top_n, top_n_pcode
    
    else:
        for i in range(n):
            bottom_n.append(fine_rank[i])
            for entry in pcode_map:
                if pcode_map[entry] == fine_rank[i]:
                    bottom_n_pcode.append(entry)
        return bottom_n, bottom_n_pcode
    
def query_specific_pcode(data,pcode):
    fine_rank = []
    pcode_map = {}
    
    for entry in data:
        fine_value = data[entry][0][1] + data[entry][1][1]
        fine_rank.append(fine_value)
        pcode_map[entry] = fine_value
        
    fine_rank = sorted(fine_rank)
    return pcode_map[pcode], len(fine_rank) - fine_rank.index(pcode_map[pcode])

def text_results(data_2014, data_2015, n):
    h_fine, h_pcode = retrieve_ranked_n(data_2014,n, top = True)
    print("Top",n,"Postcodes for Speeding Fines 2014")
    for p in range(len(h_pcode)):
        if h_pcode[p] in metro_pcode:
            print(str(p + 1) + '.\t', h_pcode[p], "$" + str(h_fine[p]) + ' - Metro')
        else:
            print(str(p + 1) + '.\t', h_pcode[p], "$" + str(h_fine[p]) + ' - Regional')

    h_fine, h_pcode = retrieve_ranked_n(data_2015,n, top = True)
    print("\nTop",n,"Postcodes for Speeding Fines 2015")
    for p in range(len(h_pcode)):
        if h_pcode[p] in metro_pcode:
            print(str(p + 1) + '.\t', h_pcode[p], "$" + str(h_fine[p]) + ' - Metro')
        else:
            print(str(p + 1) + '.\t', h_pcode[p], "$" + str(h_fine[p]) + ' - Regional')

    h_fine, h_pcode = retrieve_ranked_n(data_2014,n, top = False)
    print("\nBottom",n,"Postcodes for Speeding Fines 2014")
    for p in range(len(h_pcode)):
        if h_pcode[p] in metro_pcode:
            print(str(p + 1) + '.\t', h_pcode[p], "$" + str(h_fine[p]) + ' - Metro')
        else:
            print(str(p + 1) + '.\t', h_pcode[p], "$" + str(h_fine[p]) + ' - Regional')   

    h_fine, h_pcode = retrieve_ranked_n(data_2015,n, top = False)
    print("\nBottom",n,"Postcodes for Speeding Fines 2015")
    for p in range(len(h_pcode)):
        if h_pcode[p] in metro_pcode:
            print(str(p + 1) + '.\t', h_pcode[p], "$" + str(h_fine[p]) + ' - Metro')
        else:
            print(str(p + 1) + '.\t', h_pcode[p], "$" + str(h_fine[p]) + ' - Regional')


def circle_ranked_n(speeding_data, postcode_data, year):
    t = input("Retrieve top or bottom speeders T/B: ")
    n = int(input("How many entries (number between 1 and 300): "))
    h_fine, h_pcode = retrieve_ranked_n(speeding_data,n,top = t == "T")
    latitude = []
    longitude = []
    red = Color("red")
    blue = Color("blue")
    if t == "T":
        colors = list(red.range_to(blue,n))
    else:
        colors = list(blue.range_to(red,n))
        
    map_osm = folium.Map(location=[-33.8688, 151.2093],zoom_start = 10)
    
    for h_p in h_pcode:
        avg_lat = 0
        avg_lon = 0
        total = 0
        suburbs = []
        fine = speeding_data[h_p][0][1] + speeding_data[h_p][1][1]
        for entry in postcode_data:
            if postcode_data[entry][0] == h_p:
                avg_lat += postcode_data[entry][1]
                avg_lon += postcode_data[entry][2]
                suburbs.append(entry.title())
                total += 1
        if t == "T":
            msg = "<b>" + str(h_p) + "</b>" + "<br>Ranked <b>" + str(h_pcode.index(h_p)+1) +"</b>/596 for Speeding Fines<br>Total Fines - $<b>" + str(fine) + "</b><br>"
        else:
            msg = "<b>" + str(h_p) + "</b>" + "<br>Ranked <b>" + str(595 - h_pcode.index(h_p)+1) +"</b>/596 for Speeding Fines<br>Total Fines - $<b>" + str(fine) + "</b><br>"
        msg += ', '.join(sorted(suburbs))
        c = colors[h_pcode.index(h_p)]

        if total != 0:
            folium.CircleMarker([avg_lat/total,avg_lon/total],
                        radius=75,
                        popup=msg,
                        color=c.hex,
                        fill_color=c.hex,
                        fill = True
                       ).add_to(map_osm)   
    file = ""
    if t == "T":
        file += "Top "
    else:
        file += "Bottom "

    file += str(n) + " Postcodes for Speeding {}.html".format(year)
    
    map_osm.save(file)

    
curr_sheet, data_2014, data_2015 = retrieve_speeding_data("./speeding_stats.xlsx")
collated_data, metro_pcode = retrieve_postocde_data("./Australian_Post_Codes_Lat_Lon.xlsx")
